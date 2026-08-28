import streamlit as st
import os
import psycopg2
from psycopg2 import IntegrityError
import re
import io
import uuid
import bcrypt
import difflib
from datetime import date, datetime, timedelta
from openpyxl import Workbook, load_workbook
from collections import defaultdict


# ============================================================
# CONFIGURACIÓN
# ============================================================

st.set_page_config(
    page_title="Sistema de Gestión de Docentes",
    page_icon="📚",
    layout="wide"
)


# ============================================================
# CONEXIÓN A SUPABASE POSTGRESQL
# ============================================================

# En Render/GitHub no se usa un archivo SQLite local.
# Configura DATABASE_URL con la cadena de conexión PostgreSQL de Supabase.
DATABASE_URL = os.getenv("DATABASE_URL")

if not DATABASE_URL:
    st.error("❌ Falta configurar la variable de entorno DATABASE_URL.")
    st.stop()


class CursorPostgres:
    """Adaptador para conservar el código original con parámetros '?'."""

    def __init__(self, conexion):
        self._cursor = conexion.cursor()

    def execute(self, query, params=None):
        query_convertida = query.replace("?", "%s")
        if params is None:
            return self._cursor.execute(query_convertida)
        return self._cursor.execute(query_convertida, params)

    def fetchone(self):
        return self._cursor.fetchone()

    def fetchall(self):
        return self._cursor.fetchall()

    @property
    def rowcount(self):
        return self._cursor.rowcount

    def close(self):
        return self._cursor.close()


conexion = psycopg2.connect(DATABASE_URL, sslmode="require")
cursor = CursorPostgres(conexion)


# ============================================================
# TABLA USUARIOS (con hash y clave_texto)
# ============================================================

cursor.execute("""
CREATE TABLE IF NOT EXISTS usuarios (
    id BIGSERIAL PRIMARY KEY,
    usuario TEXT NOT NULL UNIQUE,
    clave TEXT NOT NULL,  -- hash bcrypt
    clave_texto TEXT,     -- clave en texto plano (solo para admin)
    nombre TEXT,
    rol TEXT DEFAULT 'usuario',
    activo INTEGER DEFAULT 1
)
""")
conexion.commit()

# PostgreSQL: agregar columna clave_texto si una base existente todavía no la tiene
cursor.execute("""
    ALTER TABLE usuarios
    ADD COLUMN IF NOT EXISTS clave_texto TEXT
""")
conexion.commit()

# Crear el administrador inicial solo la primera vez.
# Configura ADMIN_USER y ADMIN_PASSWORD como variables de entorno en Render.
ADMIN_USER = os.getenv("ADMIN_USER", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin")

cursor.execute("SELECT id, clave FROM usuarios WHERE usuario = ?", (ADMIN_USER,))
admin = cursor.fetchone()
if admin is None:
    hashed = bcrypt.hashpw(ADMIN_PASSWORD.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    cursor.execute(
        "INSERT INTO usuarios (usuario, clave, clave_texto, nombre, rol, activo) VALUES (?, ?, ?, ?, ?, ?)",
        (ADMIN_USER, hashed, ADMIN_PASSWORD, "Administrador", "admin", 1)
    )
    conexion.commit()
else:
    # Migración de una instalación antigua que pudiera tener una clave sin hash.
    if not admin[1].startswith("$2b$"):
        hashed = bcrypt.hashpw(admin[1].encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        cursor.execute(
            "UPDATE usuarios SET clave = ?, clave_texto = ? WHERE id = ?",
            (hashed, admin[1], admin[0])
        )
        conexion.commit()


# ============================================================
# TABLA SESIONES (tokens)
# ============================================================

cursor.execute("""
CREATE TABLE IF NOT EXISTS sesiones (
    token TEXT PRIMARY KEY,
    usuario_id INTEGER,
    fecha_creacion TEXT,
    FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
)
""")
conexion.commit()

# Limpiar tokens expirados (más de 7 días)
fecha_limite = (datetime.now() - timedelta(days=7)).isoformat()
cursor.execute("DELETE FROM sesiones WHERE fecha_creacion < ?", (fecha_limite,))
conexion.commit()


# ============================================================
# TABLA DOCENTES
# ============================================================

cursor.execute("""
CREATE TABLE IF NOT EXISTS docentes (
    id BIGSERIAL PRIMARY KEY,
    cedula TEXT NOT NULL UNIQUE,
    nombres TEXT NOT NULL,
    telefono TEXT,
    correo TEXT,
    usuario TEXT,
    clave TEXT,
    clave_modificada INTEGER DEFAULT 0,
    clave_caducada INTEGER DEFAULT 0
)
""")
conexion.commit()


# ============================================================
# TABLA CURSOS
# ============================================================

cursor.execute("""
CREATE TABLE IF NOT EXISTS productos (
    id BIGSERIAL PRIMARY KEY,
    nombre TEXT NOT NULL UNIQUE,
    descripcion TEXT,
    horas INTEGER,
    fecha_inicio TEXT,
    fecha_fin TEXT
)
""")
conexion.commit()


# ============================================================
# TABLA SERVICIOS / TRABAJOS
# ============================================================

cursor.execute("""
CREATE TABLE IF NOT EXISTS servicios (
    id BIGSERIAL PRIMARY KEY,
    docente_id INTEGER,
    descripcion TEXT,
    precio REAL NOT NULL,
    abonos REAL DEFAULT 0,
    observaciones TEXT,
    producto_id INTEGER,
    fecha_inicio TEXT,
    fecha_fin TEXT,
    FOREIGN KEY (docente_id) REFERENCES docentes(id),
    FOREIGN KEY (producto_id) REFERENCES productos(id)
)
""")
conexion.commit()


# ============================================================
# TABLA ESTADOS_CURSOS (para estados individuales por curso)
# ============================================================

cursor.execute("""
CREATE TABLE IF NOT EXISTS estados_cursos (
    id BIGSERIAL PRIMARY KEY,
    servicio_id INTEGER,
    curso_id INTEGER,
    estado TEXT,
    fecha_actualizacion TEXT,
    FOREIGN KEY (servicio_id) REFERENCES servicios(id)
)
""")
conexion.commit()


# ============================================================
# AGREGAR COLUMNAS NUEVAS SI NO EXISTEN
# ============================================================

def agregar_columna_si_no_existe(tabla, columna, tipo):
    # Los nombres se controlan desde el propio código; no provienen del usuario.
    cursor.execute(f"ALTER TABLE {tabla} ADD COLUMN IF NOT EXISTS {columna} {tipo}")
    conexion.commit()


# Docentes
agregar_columna_si_no_existe("docentes", "clave_modificada", "INTEGER DEFAULT 0")
agregar_columna_si_no_existe("docentes", "clave_caducada", "INTEGER DEFAULT 0")

# Servicios
agregar_columna_si_no_existe("servicios", "numero_pagos", "INTEGER DEFAULT 1")
agregar_columna_si_no_existe("servicios", "fecha_pago_1", "TEXT")
agregar_columna_si_no_existe("servicios", "fecha_pago_2", "TEXT")
agregar_columna_si_no_existe("servicios", "fecha_pago_3", "TEXT")
agregar_columna_si_no_existe("servicios", "monto_pago_1", "REAL DEFAULT 0")
agregar_columna_si_no_existe("servicios", "monto_pago_2", "REAL DEFAULT 0")
agregar_columna_si_no_existe("servicios", "monto_pago_3", "REAL DEFAULT 0")
agregar_columna_si_no_existe("servicios", "fecha_estimada_1", "TEXT")
agregar_columna_si_no_existe("servicios", "fecha_estimada_2", "TEXT")
agregar_columna_si_no_existe("servicios", "fecha_estimada_3", "TEXT")
agregar_columna_si_no_existe("servicios", "tipo_trabajo", "TEXT DEFAULT 'CURSO'")
agregar_columna_si_no_existe("servicios", "promo_cantidad", "INTEGER DEFAULT 0")
agregar_columna_si_no_existe("servicios", "promo_curso_1", "INTEGER")
agregar_columna_si_no_existe("servicios", "promo_curso_2", "INTEGER")
agregar_columna_si_no_existe("servicios", "promo_curso_3", "INTEGER")
agregar_columna_si_no_existe("servicios", "promo_curso_4", "INTEGER")
agregar_columna_si_no_existe("servicios", "promo_curso_5", "INTEGER")
agregar_columna_si_no_existe("servicios", "fecha_inicio_trabajo", "TEXT")


# ============================================================
# ELIMINAR ÍNDICE ÚNICO ANTERIOR SI EXISTE
# ============================================================

cursor.execute("DROP INDEX IF EXISTS indice_docente_curso_unico")
conexion.commit()


# ============================================================
# FUNCIONES AUXILIARES DE VALIDACIÓN DE CLAVE SIMILAR
# ============================================================

def clave_similar_a_usuario(usuario, clave, umbral=0.8):
    """
    Retorna True si la clave es demasiado similar al usuario.
    Se considera similar si:
    - son iguales
    - el usuario está contenido en la clave (o viceversa)
    - la similitud de secuencia (SequenceMatcher) es mayor a umbral (0.8)
    """
    usuario = usuario.lower()
    clave = clave.lower()
    if usuario == clave:
        return True
    if usuario in clave or clave in usuario:
        return True
    # Calcular similitud con SequenceMatcher
    ratio = difflib.SequenceMatcher(None, usuario, clave).ratio()
    return ratio > umbral


# ============================================================
# VALIDACIONES
# ============================================================

def validar_cedula_ecuatoriana(cedula):
    cedula = str(cedula).strip()
    if len(cedula) != 10:
        return False, "La cédula debe tener exactamente 10 números."
    if not cedula.isdigit():
        return False, "La cédula solo puede contener números."
    if len(set(cedula)) == 1:
        return False, "El número de cédula no puede estar formado por un mismo número repetido."
    provincia = int(cedula[:2])
    if provincia < 1 or provincia > 24:
        return False, "La cédula no tiene un código provincial válido de Ecuador."
    tercer_digito = int(cedula[2])
    if tercer_digito > 5:
        return False, "El número de cédula ecuatoriana no es válido."
    coeficientes = [2, 1, 2, 1, 2, 1, 2, 1, 2]
    suma = 0
    for i in range(9):
        resultado = int(cedula[i]) * coeficientes[i]
        if resultado >= 10:
            resultado -= 9
        suma += resultado
    residuo = suma % 10
    digito_verificador = 0 if residuo == 0 else 10 - residuo
    if digito_verificador != int(cedula[9]):
        return False, "El número de cédula ecuatoriana no es válido."
    return True, ""


def validar_nombres(nombres):
    nombres = str(nombres).strip()
    if len(nombres) < 4:
        return False, "Los nombres deben tener como mínimo 4 caracteres."
    patron = r"^[A-Za-zÁÉÍÓÚáéíóúÑñÜü ]+$"
    if not re.fullmatch(patron, nombres):
        return False, "Los nombres solo pueden contener letras y espacios."
    return True, ""


def validar_telefono(telefono):
    telefono = str(telefono).strip()
    if not telefono.isdigit():
        return False, "El teléfono solo puede contener números."
    if len(telefono) != 10:
        return False, "El teléfono debe tener exactamente 10 números."
    return True, ""


def validar_correo(correo):
    correo = str(correo).strip()
    patron = r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$"
    if not re.fullmatch(patron, correo):
        return False, "Ingresa un correo electrónico válido."
    return True, ""


def cedula_ya_registrada(cedula, id_excluir=None):
    if id_excluir is None:
        cursor.execute("SELECT id FROM docentes WHERE cedula = ?", (cedula,))
    else:
        cursor.execute("SELECT id FROM docentes WHERE cedula = ? AND id != ?", (cedula, id_excluir))
    return cursor.fetchone() is not None


def docente_tiene_curso(docente_id, curso_id):
    cursor.execute(
        "SELECT id FROM servicios WHERE docente_id = ? AND producto_id = ? LIMIT 1",
        (docente_id, curso_id)
    )
    if cursor.fetchone() is not None:
        return True
    cursor.execute(
        """
        SELECT id FROM servicios 
        WHERE docente_id = ? AND (
            promo_curso_1 = ? OR promo_curso_2 = ? OR promo_curso_3 = ? 
            OR promo_curso_4 = ? OR promo_curso_5 = ?
        ) LIMIT 1
        """,
        (docente_id, curso_id, curso_id, curso_id, curso_id, curso_id)
    )
    return cursor.fetchone() is not None


def cursos_repetidos(lista_cursos):
    lista_validos = [curso for curso in lista_cursos if curso is not None]
    return len(lista_validos) != len(set(lista_validos))


def cantidad_servicios_docente(docente_id):
    cursor.execute("SELECT COUNT(*) FROM servicios WHERE docente_id = ?", (docente_id,))
    return cursor.fetchone()[0]


def cantidad_servicios_curso(curso_id):
    cursor.execute(
        """
        SELECT COUNT(*) FROM servicios 
        WHERE producto_id = ? OR promo_curso_1 = ? OR promo_curso_2 = ? 
        OR promo_curso_3 = ? OR promo_curso_4 = ? OR promo_curso_5 = ?
        """,
        (curso_id, curso_id, curso_id, curso_id, curso_id, curso_id)
    )
    return cursor.fetchone()[0]


def crear_plantilla_excel():
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Docentes"
    encabezados = ["Número de cédula", "Nombres", "Teléfono", "Correo", "Usuario", "Clave"]
    for columna, encabezado in enumerate(encabezados, start=1):
        hoja.cell(row=1, column=columna).value = encabezado
    for fila in range(2, 1002):
        hoja.cell(row=fila, column=1).number_format = "@"
        hoja.cell(row=fila, column=3).number_format = "@"
    anchos = {"A": 20, "B": 30, "C": 18, "D": 35, "E": 20, "F": 20}
    for columna, ancho in anchos.items():
        hoja.column_dimensions[columna].width = ancho
    archivo = io.BytesIO()
    libro.save(archivo)
    archivo.seek(0)
    return archivo


# ============================================================
# FUNCIONES PARA GESTIONAR ESTADOS DE CURSOS
# ============================================================

# Estados posibles
ESTADOS_POSIBLES = ["Matriculado", "En proceso", "Terminado", "Entregado", "No salió", "Cancelado", "Próxima matrícula"]

def obtener_cursos_servicio(servicio_id):
    """Devuelve una lista de (curso_id, nombre_curso, horas) para un servicio dado."""
    cursor.execute("SELECT tipo_trabajo, producto_id, promo_curso_1, promo_curso_2, promo_curso_3, promo_curso_4, promo_curso_5 FROM servicios WHERE id = ?", (servicio_id,))
    datos = cursor.fetchone()
    if not datos:
        return []
    tipo_trabajo = datos[0]
    if tipo_trabajo == "PROMO":
        cursos_ids = []
        for i in range(1, 6):
            curso_id = datos[i]
            if curso_id is not None:
                cursos_ids.append(curso_id)
    else:
        cursos_ids = [datos[1]]  # producto_id
    cursos = []
    for cid in cursos_ids:
        cursor.execute("SELECT nombre, horas FROM productos WHERE id = ?", (cid,))
        resultado = cursor.fetchone()
        if resultado:
            cursos.append((cid, resultado[0], resultado[1]))  # (id, nombre, horas)
    return cursos


def obtener_estados_servicio(servicio_id):
    """Devuelve un diccionario {curso_id: estado} para un servicio."""
    cursor.execute("SELECT curso_id, estado FROM estados_cursos WHERE servicio_id = ?", (servicio_id,))
    estados = {}
    for cid, estado in cursor.fetchall():
        estados[cid] = estado
    return estados


def actualizar_estado_curso(servicio_id, curso_id, nuevo_estado):
    """Actualiza el estado de un curso en un servicio."""
    if nuevo_estado not in ESTADOS_POSIBLES:
        return False
    fecha = datetime.now().isoformat()
    cursor.execute(
        "UPDATE estados_cursos SET estado = ?, fecha_actualizacion = ? WHERE servicio_id = ? AND curso_id = ?",
        (nuevo_estado, fecha, servicio_id, curso_id)
    )
    if cursor.rowcount == 0:
        # Si no existía, insertar (por si acaso)
        cursor.execute(
            "INSERT INTO estados_cursos (servicio_id, curso_id, estado, fecha_actualizacion) VALUES (?, ?, ?, ?)",
            (servicio_id, curso_id, nuevo_estado, fecha)
        )
    conexion.commit()
    return True


def inicializar_estados(servicio_id, cursos_ids):
    """Inserta estados iniciales para un nuevo servicio."""
    fecha = datetime.now().isoformat()
    for cid in cursos_ids:
        cursor.execute(
            "INSERT INTO estados_cursos (servicio_id, curso_id, estado, fecha_actualizacion) VALUES (?, ?, ?, ?)",
            (servicio_id, cid, "Matriculado", fecha)
        )
    conexion.commit()


def eliminar_estados(servicio_id):
    cursor.execute("DELETE FROM estados_cursos WHERE servicio_id = ?", (servicio_id,))
    conexion.commit()


# ============================================================
# INICIALIZAR ESTADO DE SESIÓN
# ============================================================

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.user_id = None
    st.session_state.user_name = None
    st.session_state.user_rol = None

if "opcion_menu" not in st.session_state:
    st.session_state.opcion_menu = "🏠 Inicio"

if "docentes_abiertos" not in st.session_state:
    st.session_state.docentes_abiertos = set()

if "docente_editando" not in st.session_state:
    st.session_state.docente_editando = None

if "docente_eliminando" not in st.session_state:
    st.session_state.docente_eliminando = None

if "curso_editando" not in st.session_state:
    st.session_state.curso_editando = None

if "curso_eliminando" not in st.session_state:
    st.session_state.curso_eliminando = None

if "servicio_editando" not in st.session_state:
    st.session_state.servicio_editando = None

if "servicio_eliminando" not in st.session_state:
    st.session_state.servicio_eliminando = None

if "trabajos_abiertos" not in st.session_state:
    st.session_state.trabajos_abiertos = set()


# ============================================================
# FUNCIONES DE AUTENTICACIÓN Y USUARIOS (con bcrypt y tokens)
# ============================================================

def verificar_clave(usuario_id, clave_plana):
    cursor.execute("SELECT clave FROM usuarios WHERE id = ?", (usuario_id,))
    resultado = cursor.fetchone()
    if resultado is None:
        return False
    try:
        return bcrypt.checkpw(clave_plana.encode('utf-8'), resultado[0].encode('utf-8'))
    except ValueError:
        return False


def autenticar_usuario(usuario, clave):
    cursor.execute("SELECT id, nombre, rol, clave FROM usuarios WHERE usuario = ? AND activo = 1", (usuario.strip(),))
    resultado = cursor.fetchone()
    if resultado is None:
        return None, None, None
    user_id, user_name, user_rol, hash_clave = resultado
    try:
        if bcrypt.checkpw(clave.encode('utf-8'), hash_clave.encode('utf-8')):
            return user_id, user_name, user_rol
    except ValueError:
        pass
    return None, None, None


def generar_token(usuario_id):
    token = str(uuid.uuid4())
    fecha_creacion = datetime.now().isoformat()
    cursor.execute("DELETE FROM sesiones WHERE usuario_id = ?", (usuario_id,))
    cursor.execute("INSERT INTO sesiones (token, usuario_id, fecha_creacion) VALUES (?, ?, ?)", (token, usuario_id, fecha_creacion))
    conexion.commit()
    return token


def validar_token(token):
    cursor.execute("SELECT usuario_id, fecha_creacion FROM sesiones WHERE token = ?", (token,))
    resultado = cursor.fetchone()
    if resultado is None:
        return None
    usuario_id, fecha_creacion = resultado
    fecha_creacion_dt = datetime.fromisoformat(fecha_creacion)
    if datetime.now() - fecha_creacion_dt > timedelta(days=7):
        cursor.execute("DELETE FROM sesiones WHERE token = ?", (token,))
        conexion.commit()
        return None
    cursor.execute("SELECT nombre, rol FROM usuarios WHERE id = ? AND activo = 1", (usuario_id,))
    user_data = cursor.fetchone()
    if user_data is None:
        return None
    return usuario_id, user_data[0], user_data[1]


def eliminar_token(token):
    cursor.execute("DELETE FROM sesiones WHERE token = ?", (token,))
    conexion.commit()


def login():
    st.title("🔐 Inicio de sesión")
    st.markdown("Ingresa tus credenciales para acceder al sistema.")
    with st.form("form_login"):
        usuario = st.text_input("Usuario")
        clave = st.text_input("Clave", type="password")
        enviar = st.form_submit_button("Iniciar sesión")

        if enviar:
            if usuario and clave:
                user_id, user_name, user_rol = autenticar_usuario(usuario, clave)
                if user_id:
                    token = generar_token(user_id)
                    st.query_params["token"] = token
                    st.session_state.logged_in = True
                    st.session_state.user_id = user_id
                    st.session_state.user_name = user_name or usuario
                    st.session_state.user_rol = user_rol
                    st.session_state.opcion_menu = "🏠 Inicio"
                    st.success(f"✅ Bienvenido, {st.session_state.user_name}!")
                    st.rerun()
                else:
                    st.error("❌ Usuario o clave incorrectos, o cuenta inactiva.")
            else:
                st.warning("⚠️ Por favor, completa ambos campos.")


def logout():
    token = st.query_params.get("token", None)
    if token:
        eliminar_token(token)
    st.query_params.clear()
    st.session_state.logged_in = False
    st.session_state.user_id = None
    st.session_state.user_name = None
    st.session_state.user_rol = None
    st.session_state.opcion_menu = "🏠 Inicio"
    st.rerun()


def cambiar_clave(usuario_id, nueva_clave_plana):
    # Obtener el nombre de usuario para validar similitud
    cursor.execute("SELECT usuario FROM usuarios WHERE id = ?", (usuario_id,))
    row = cursor.fetchone()
    if row is None:
        return False, "Usuario no encontrado."
    usuario = row[0]
    if clave_similar_a_usuario(usuario, nueva_clave_plana):
        return False, "La nueva clave es demasiado similar al nombre de usuario. Elige una clave más diferente."
    hashed = bcrypt.hashpw(nueva_clave_plana.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    cursor.execute(
        "UPDATE usuarios SET clave = ?, clave_texto = ? WHERE id = ?",
        (hashed, nueva_clave_plana, usuario_id)
    )
    conexion.commit()
    return True, "Clave actualizada correctamente."


def crear_usuario(usuario, clave_plana, nombre, rol="usuario", activo=1):
    if clave_similar_a_usuario(usuario, clave_plana):
        return False, "La clave es demasiado similar al nombre de usuario. Elige una clave más diferente."
    try:
        hashed = bcrypt.hashpw(clave_plana.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        cursor.execute(
            "INSERT INTO usuarios (usuario, clave, clave_texto, nombre, rol, activo) VALUES (?, ?, ?, ?, ?, ?)",
            (usuario.strip(), hashed, clave_plana, nombre.strip(), rol, activo)
        )
        conexion.commit()
        return True, "Usuario creado correctamente."
    except IntegrityError:
        conexion.rollback()
        return False, "El nombre de usuario ya existe."


def actualizar_usuario(id_usuario, usuario=None, nombre=None, rol=None, activo=None):
    campos = []
    valores = []
    if usuario is not None:
        campos.append("usuario = ?")
        valores.append(usuario.strip())
    if nombre is not None:
        campos.append("nombre = ?")
        valores.append(nombre.strip())
    if rol is not None:
        campos.append("rol = ?")
        valores.append(rol)
    if activo is not None:
        campos.append("activo = ?")
        valores.append(1 if activo else 0)
    if not campos:
        return False
    valores.append(id_usuario)
    query = f"UPDATE usuarios SET {', '.join(campos)} WHERE id = ?"
    cursor.execute(query, valores)
    conexion.commit()
    return cursor.rowcount > 0


def obtener_usuarios():
    cursor.execute("SELECT id, usuario, nombre, rol, activo, clave_texto FROM usuarios ORDER BY id")
    return cursor.fetchall()


def es_admin():
    return st.session_state.user_rol == "admin"


# ============================================================
# VERIFICAR SESIÓN AL INICIO
# ============================================================

token_url = st.query_params.get("token", None)
if token_url and not st.session_state.logged_in:
    user_data = validar_token(token_url)
    if user_data:
        user_id, user_name, user_rol = user_data
        st.session_state.logged_in = True
        st.session_state.user_id = user_id
        st.session_state.user_name = user_name
        st.session_state.user_rol = user_rol
        st.session_state.opcion_menu = "🏠 Inicio"
    else:
        st.query_params.clear()
        st.rerun()


# ============================================================
# PANTALLA DE LOGIN O APLICACIÓN PRINCIPAL
# ============================================================

if not st.session_state.logged_in:
    login()
    st.stop()


# ============================================================
# MENÚ PRINCIPAL
# ============================================================

st.sidebar.markdown(f"**👤 Usuario:** {st.session_state.user_name}")
st.sidebar.markdown(f"**🔑 Rol:** {st.session_state.user_rol.capitalize()}")
if st.sidebar.button("🚪 Cerrar sesión", use_container_width=True):
    logout()
st.sidebar.divider()

if es_admin():
    if st.sidebar.button("👥 Administrar usuarios", use_container_width=True, type="primary" if st.session_state.opcion_menu == "👥 Administrar usuarios" else "secondary"):
        st.session_state.opcion_menu = "👥 Administrar usuarios"
        st.rerun()
    st.sidebar.divider()

st.sidebar.markdown("## 📋 Navegación")
opciones_menu = [
    "🏠 Inicio",
    "👨‍🏫 Agregar docente",
    "📥 Importar docentes",
    "👥 Gestionar docentes",
    "📚 Lista de cursos",
    "🆕 Nuevo trabajo",
    "📋 Trabajos en proceso"
]

for opcion in opciones_menu:
    if st.session_state.opcion_menu == opcion:
        if st.sidebar.button(opcion, use_container_width=True, type="primary"):
            st.session_state.opcion_menu = opcion
            st.rerun()
    else:
        if st.sidebar.button(opcion, use_container_width=True):
            st.session_state.opcion_menu = opcion
            st.rerun()

st.sidebar.divider()
opcion = st.session_state.opcion_menu
st.title("📚 Sistema de Gestión de Docentes y Cursos")


# ============================================================
# INICIO
# ============================================================

if opcion == "🏠 Inicio":
    st.header("🏠 Inicio")
    cursor.execute("SELECT COUNT(*) FROM docentes")
    cantidad_docentes = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM productos")
    cantidad_cursos = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM servicios")
    cantidad_trabajos = cursor.fetchone()[0]
    cursor.execute("SELECT COALESCE(SUM(precio), 0) FROM servicios")
    total = cursor.fetchone()[0]
    cursor.execute("SELECT COALESCE(SUM(abonos), 0) FROM servicios")
    abonos_total = cursor.fetchone()[0]
    saldo = total - abonos_total

    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.metric("👨‍🏫 Docentes", cantidad_docentes)
    with col2:
        st.metric("📚 Cursos", cantidad_cursos)
    with col3:
        st.metric("🆕 Trabajos", cantidad_trabajos)
    with col4:
        st.metric("💰 Total", f"${total:.2f}")
    with col5:
        st.metric("💵 Saldo", f"${saldo:.2f}")
    st.divider()
    st.info("Utiliza el menú de la izquierda para administrar docentes, cursos y trabajos.")


# ============================================================
# AGREGAR DOCENTE
# ============================================================

elif opcion == "👨‍🏫 Agregar docente":
    st.header("👨‍🏫 Agregar docente")
    with st.form("formulario_agregar_docente"):
        cedula = st.text_input("Número de cédula", max_chars=10)
        nombres = st.text_input("Nombres")
        telefono = st.text_input("Teléfono", max_chars=10)
        correo = st.text_input("Correo")
        usuario = st.text_input("Usuario")
        clave = st.text_input("Clave")
        guardar = st.form_submit_button("💾 Guardar docente")

        if guardar:
            valida, mensaje = validar_cedula_ecuatoriana(cedula)
            if not valida:
                st.error(mensaje)
            elif cedula_ya_registrada(cedula.strip()):
                st.error("Número de cédula ya registrado.")
            else:
                valida, mensaje = validar_nombres(nombres)
                if not valida:
                    st.error(mensaje)
                else:
                    valida, mensaje = validar_telefono(telefono)
                    if not valida:
                        st.error(mensaje)
                    elif telefono.strip() == cedula.strip():
                        st.error("El número de teléfono no puede ser igual al número de cédula.")
                    else:
                        valida, mensaje = validar_correo(correo)
                        if not valida:
                            st.error(mensaje)
                        else:
                            try:
                                cursor.execute(
                                    """
                                    INSERT INTO docentes (cedula, nombres, telefono, correo, usuario, clave, clave_modificada, clave_caducada)
                                    VALUES (?, ?, ?, ?, ?, ?, 0, 0)
                                    """,
                                    (cedula.strip(), nombres.strip(), telefono.strip(), correo.strip(), usuario.strip(), clave.strip())
                                )
                                conexion.commit()
                                st.success("✅ Docente registrado correctamente.")
                            except IntegrityError:
                                conexion.rollback()
                                st.error("Número de cédula ya registrado.")


# ============================================================
# IMPORTAR DOCENTES
# ============================================================

elif opcion == "📥 Importar docentes":
    st.header("📥 Importar docentes desde Excel")
    plantilla = crear_plantilla_excel()
    st.download_button(
        label="📥 Descargar plantilla Excel",
        data=plantilla,
        file_name="plantilla_docentes.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.divider()
    archivo = st.file_uploader("Selecciona tu archivo Excel", type=["xlsx"])
    if archivo is not None:
        try:
            libro = load_workbook(archivo, data_only=True)
            hoja = libro.active
            filas = list(hoja.iter_rows(values_only=True))
            if len(filas) < 2:
                st.error("El archivo no contiene docentes.")
            else:
                encabezados = ["Número de cédula", "Nombres", "Teléfono", "Correo", "Usuario", "Clave"]
                encabezados_excel = [str(valor).strip() if valor is not None else "" for valor in filas[0]]
                if encabezados_excel[:6] != encabezados:
                    st.error("❌ El formato del Excel no es correcto.")
                else:
                    importados = 0
                    rechazados = 0
                    errores = []
                    cedulas_excel = set()
                    for numero_fila, fila in enumerate(filas[1:], start=2):
                        valores = list(fila[:6])
                        while len(valores) < 6:
                            valores.append("")
                        valor_cedula = valores[0]
                        if valor_cedula is None:
                            cedula = ""
                        elif isinstance(valor_cedula, int):
                            cedula = str(valor_cedula).zfill(10)
                        elif isinstance(valor_cedula, float):
                            cedula = str(int(valor_cedula)).zfill(10)
                        else:
                            cedula = str(valor_cedula).strip()
                            if cedula.endswith(".0"):
                                cedula = cedula[:-2]
                            if cedula.isdigit():
                                cedula = cedula.zfill(10)

                        nombres = "" if valores[1] is None else str(valores[1]).strip()

                        if valores[2] is None:
                            telefono = ""
                        elif isinstance(valores[2], int):
                            telefono = str(valores[2]).zfill(10)
                        elif isinstance(valores[2], float):
                            telefono = str(int(valores[2])).zfill(10)
                        else:
                            telefono = str(valores[2]).strip()

                        correo = "" if valores[3] is None else str(valores[3]).strip()
                        usuario = "" if valores[4] is None else str(valores[4]).strip()
                        clave = "" if valores[5] is None else str(valores[5]).strip()

                        if cedula == "" and nombres == "" and telefono == "" and correo == "":
                            continue
                        errores_fila = []
                        valida, mensaje = validar_cedula_ecuatoriana(cedula)
                        if not valida:
                            errores_fila.append(mensaje)
                        else:
                            if cedula in cedulas_excel:
                                errores_fila.append("Número de cédula repetido dentro del archivo.")
                            else:
                                cedulas_excel.add(cedula)
                            if cedula_ya_registrada(cedula):
                                errores_fila.append("Número de cédula ya registrado en el sistema.")

                        valida, mensaje = validar_nombres(nombres)
                        if not valida:
                            errores_fila.append(mensaje)
                        valida, mensaje = validar_telefono(telefono)
                        if not valida:
                            errores_fila.append(mensaje)
                        elif telefono == cedula:
                            errores_fila.append("El teléfono no puede ser igual a la cédula.")
                        valida, mensaje = validar_correo(correo)
                        if not valida:
                            errores_fila.append(mensaje)

                        if errores_fila:
                            rechazados += 1
                            errores.append((numero_fila, cedula, errores_fila))
                        else:
                            try:
                                cursor.execute(
                                    """
                                    INSERT INTO docentes (cedula, nombres, telefono, correo, usuario, clave, clave_modificada, clave_caducada)
                                    VALUES (?, ?, ?, ?, ?, ?, 0, 0)
                                    """,
                                    (cedula, nombres, telefono, correo, usuario, clave)
                                )
                                conexion.commit()
                                importados += 1
                            except IntegrityError:
                                conexion.rollback()
                                rechazados += 1
                                errores.append((numero_fila, cedula, ["Número de cédula ya registrado."]))
                    st.divider()
                    st.subheader("📊 Resultado")
                    st.success(f"✅ {importados} docente(s) importado(s) correctamente.")
                    if rechazados > 0:
                        st.warning(f"⚠️ {rechazados} docente(s) no fueron importados.")
                        for fila, cedula, mensajes in errores:
                            st.error(f"Fila {fila} - Cédula {cedula}")
                            for mensaje in mensajes:
                                st.write(f"• {mensaje}")
                    else:
                        st.success("🎉 Todos los docentes fueron importados.")
        except Exception as e:
            st.error("❌ No se pudo leer el archivo Excel.")
            st.caption(f"Detalle: {e}")


# ============================================================
# GESTIONAR DOCENTES (con paginación)
# ============================================================

elif opcion == "👥 Gestionar docentes":
    st.header("👥 Gestionar docentes")

    # Inicializar estado de paginación
    if "pagina_docentes" not in st.session_state:
        st.session_state.pagina_docentes = 1
    if "docentes_por_pagina" not in st.session_state:
        st.session_state.docentes_por_pagina = 10

    busqueda = st.text_input("🔎 Buscar docente", placeholder="Nombre, cédula, teléfono, correo o usuario...")

    # Botones para abrir/colapsar todos
    col1, col2 = st.columns(2)
    with col1:
        if st.button("📂 Abrir todos", use_container_width=True):
            cursor.execute("SELECT id FROM docentes")
            ids = cursor.fetchall()
            st.session_state.docentes_abiertos = {fila[0] for fila in ids}
            st.rerun()
    with col2:
        if st.button("📁 Colapsar todos", use_container_width=True):
            st.session_state.docentes_abiertos = set()
            st.session_state.docente_editando = None
            st.session_state.docente_eliminando = None
            st.rerun()

    # Obtener datos según búsqueda
    if busqueda.strip() == "":
        cursor.execute(
            "SELECT id, cedula, nombres, telefono, correo, usuario, clave, clave_modificada, clave_caducada FROM docentes ORDER BY nombres"
        )
    else:
        texto = f"%{busqueda.strip()}%"
        cursor.execute(
            """
            SELECT id, cedula, nombres, telefono, correo, usuario, clave, clave_modificada, clave_caducada
            FROM docentes 
            WHERE nombres LIKE ? OR cedula LIKE ? OR telefono LIKE ? OR correo LIKE ? OR usuario LIKE ?
            ORDER BY nombres
            """,
            (texto, texto, texto, texto, texto)
        )
    docentes = cursor.fetchall()
    total_docentes = len(docentes)

    # Si hay búsqueda, reiniciar a página 1
    if busqueda.strip() != "" and st.session_state.pagina_docentes != 1:
        st.session_state.pagina_docentes = 1

    # Calcular paginación
    por_pagina = st.session_state.docentes_por_pagina
    total_paginas = max(1, (total_docentes + por_pagina - 1) // por_pagina)
    if st.session_state.pagina_docentes > total_paginas:
        st.session_state.pagina_docentes = total_paginas
    pagina_actual = st.session_state.pagina_docentes
    inicio = (pagina_actual - 1) * por_pagina
    fin = min(inicio + por_pagina, total_docentes)
    docentes_pagina = docentes[inicio:fin]

    # Función para mostrar controles de paginación (se usa arriba y abajo)
    def mostrar_controles_paginacion(posicion):
        """posicion: 'top' o 'bottom' para claves únicas"""
        col_ant, col_pag, col_sig, col_sel = st.columns([1, 2, 1, 2])
        with col_ant:
            if st.button("◀ Anterior", key=f"prev_{posicion}", disabled=(st.session_state.pagina_docentes <= 1)):
                st.session_state.pagina_docentes -= 1
                st.rerun()
        with col_pag:
            st.write(f"Página {st.session_state.pagina_docentes} de {total_paginas}")
        with col_sig:
            if st.button("Siguiente ▶", key=f"next_{posicion}", disabled=(st.session_state.pagina_docentes >= total_paginas)):
                st.session_state.pagina_docentes += 1
                st.rerun()
        with col_sel:
            nueva_por_pagina = st.selectbox(
                "Ver por página",
                [5, 10, 25, 50],
                index=[5, 10, 25, 50].index(st.session_state.docentes_por_pagina),
                key=f"select_por_pagina_{posicion}",
                label_visibility="collapsed"
            )
            if nueva_por_pagina != st.session_state.docentes_por_pagina:
                st.session_state.docentes_por_pagina = nueva_por_pagina
                st.session_state.pagina_docentes = 1
                st.rerun()

    # Mostrar info y controles SUPERIOR
    if total_docentes > 0:
        st.write(f"**Mostrando {inicio + 1} - {fin} de {total_docentes} docentes**")
    else:
        st.write("**No se encontraron docentes**")

    mostrar_controles_paginacion("top")
    st.divider()

    # Mostrar docentes de la página actual
    if not docentes_pagina:
        st.info("📁 No hay docentes en esta página.")
    else:
        for docente in docentes_pagina:
            docente_id = docente[0]
            cedula = docente[1]
            nombres = docente[2]
            telefono = docente[3]
            correo = docente[4]
            usuario = docente[5]
            clave = docente[6]
            clave_modificada = docente[7]
            clave_caducada = docente[8]

            abierto = docente_id in st.session_state.docentes_abiertos
            col1, col2 = st.columns([5, 1])
            with col1:
                st.markdown(f"### 👨‍🏫 {nombres}")
                st.caption(f"Cédula: {cedula}")
            with col2:
                texto_boton = "📁 Ocultar" if abierto else "📂 Mostrar"
                if st.button(texto_boton, key=f"mostrar_{docente_id}", use_container_width=True):
                    if abierto:
                        st.session_state.docentes_abiertos.remove(docente_id)
                    else:
                        st.session_state.docentes_abiertos.add(docente_id)
                    st.rerun()
            if abierto:
                with st.container(border=True):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**Cédula:** {cedula}")
                        st.write(f"**Nombres:** {nombres}")
                        st.write(f"**Teléfono:** {telefono}")
                    with col2:
                        st.write(f"**Correo:** {correo}")
                        st.write(f"**Usuario:** {usuario}")
                        if clave is None or clave.strip() == "":
                            color = "red"
                            estado = "Vacía"
                        elif clave_caducada == 1:
                            color = "red"
                            estado = "Caducada"
                        elif clave_modificada == 1:
                            color = "orange"
                            estado = "Modificada"
                        else:
                            color = "green"
                            estado = "Original"
                        st.markdown(
                            f"**Clave:** <span style='color:{color}; font-weight:bold;'>{clave}</span> "
                            f"<span style='font-size:0.8em; color:gray;'>({estado})</span>",
                            unsafe_allow_html=True
                        )
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("✏️ Editar", key=f"editar_{docente_id}", use_container_width=True):
                            st.session_state.docente_editando = docente_id
                            st.session_state.docente_eliminando = None
                            st.rerun()
                    with col2:
                        if st.button("🗑️ Eliminar", key=f"eliminar_{docente_id}", use_container_width=True):
                            st.session_state.docente_eliminando = docente_id
                            st.session_state.docente_editando = None
                            st.rerun()

                    if st.session_state.docente_editando == docente_id:
                        st.divider()
                        st.subheader("✏️ Editar docente")
                        with st.form(f"editar_docente_{docente_id}"):
                            nueva_cedula = st.text_input("Número de cédula", value=cedula, max_chars=10)
                            nuevos_nombres = st.text_input("Nombres", value=nombres)
                            nuevo_telefono = st.text_input("Teléfono", value=telefono, max_chars=10)
                            nuevo_correo = st.text_input("Correo", value=correo)
                            nuevo_usuario = st.text_input("Usuario", value=usuario)
                            nueva_clave = st.text_input("Clave", value=clave)
                            caducada = st.checkbox("Clave caducada", value=bool(clave_caducada))
                            guardar = st.form_submit_button("💾 Guardar cambios")

                            if guardar:
                                valida, mensaje = validar_cedula_ecuatoriana(nueva_cedula)
                                if not valida:
                                    st.error(mensaje)
                                elif cedula_ya_registrada(nueva_cedula, docente_id):
                                    st.error("Número de cédula ya registrado.")
                                else:
                                    valida, mensaje = validar_nombres(nuevos_nombres)
                                    if not valida:
                                        st.error(mensaje)
                                    else:
                                        valida, mensaje = validar_telefono(nuevo_telefono)
                                        if not valida:
                                            st.error(mensaje)
                                        elif nuevo_telefono == nueva_cedula:
                                            st.error("El teléfono no puede ser igual a la cédula.")
                                        else:
                                            valida, mensaje = validar_correo(nuevo_correo)
                                            if not valida:
                                                st.error(mensaje)
                                            else:
                                                if nueva_clave != clave:
                                                    nuevo_flag_mod = 1
                                                else:
                                                    nuevo_flag_mod = clave_modificada
                                                nuevo_flag_cad = 1 if caducada else 0
                                                cursor.execute(
                                                    """
                                                    UPDATE docentes 
                                                    SET cedula = ?, nombres = ?, telefono = ?, correo = ?, usuario = ?, 
                                                        clave = ?, clave_modificada = ?, clave_caducada = ?
                                                    WHERE id = ?
                                                    """,
                                                    (
                                                        nueva_cedula.strip(),
                                                        nuevos_nombres.strip(),
                                                        nuevo_telefono.strip(),
                                                        nuevo_correo.strip(),
                                                        nuevo_usuario.strip(),
                                                        nueva_clave.strip(),
                                                        nuevo_flag_mod,
                                                        nuevo_flag_cad,
                                                        docente_id
                                                    )
                                                )
                                                conexion.commit()
                                                st.session_state.docente_editando = None
                                                st.success("✅ Docente actualizado correctamente.")
                                                st.rerun()

                    if st.session_state.docente_eliminando == docente_id:
                        cantidad = cantidad_servicios_docente(docente_id)
                        if cantidad > 0:
                            st.error("🚫 No se puede eliminar este docente.")
                            st.warning(f"Tiene {cantidad} trabajo(s) asociado(s).")
                        else:
                            confirmar = st.checkbox("Confirmo que deseo eliminar este docente.", key=f"confirmar_{docente_id}")
                            if confirmar:
                                if st.button("🗑️ Confirmar eliminación", key=f"confirmar_eliminar_{docente_id}"):
                                    cursor.execute("DELETE FROM docentes WHERE id = ?", (docente_id,))
                                    conexion.commit()
                                    st.session_state.docente_eliminando = None
                                    st.session_state.docentes_abiertos.discard(docente_id)
                                    st.success("✅ Docente eliminado correctamente.")
                                    st.rerun()
            st.divider()

        # Controles de paginación INFERIOR
        st.divider()
        mostrar_controles_paginacion("bottom")


# ============================================================
# LISTA DE CURSOS
# ============================================================

elif opcion == "📚 Lista de cursos":
    st.header("📚 Lista de cursos")
    st.write("Aquí puedes registrar y administrar los cursos que ofreces.")
    st.divider()
    pestaña1, pestaña2 = st.tabs(["➕ Agregar curso", "📋 Gestionar cursos"])
    
    with pestaña1:
        st.subheader("➕ Agregar curso")
        with st.form("form_agregar_curso"):
            nombre_curso = st.text_input("Nombre del curso")
            horas_curso = st.selectbox("Número de horas", list(range(40, 331)), format_func=lambda x: f"{x} horas")
            guardar = st.form_submit_button("💾 Guardar curso")
            if guardar:
                if nombre_curso.strip() == "":
                    st.error("Debes ingresar el nombre del curso.")
                else:
                    try:
                        cursor.execute("INSERT INTO productos (nombre, horas) VALUES (?, ?)", (nombre_curso.strip(), horas_curso))
                        conexion.commit()
                        st.success("✅ Curso registrado correctamente.")
                    except IntegrityError:
                        conexion.rollback()
                        st.error("⚠️ Ese curso ya está registrado.")
                        
    with pestaña2:
        st.subheader("📋 Cursos registrados")
        cursor.execute("SELECT id, nombre, horas FROM productos ORDER BY nombre")
        cursos = cursor.fetchall()
        if not cursos:
            st.info("Todavía no hay cursos registrados.")
        else:
            for curso in cursos:
                curso_id = curso[0]
                nombre = curso[1]
                horas = curso[2]
                with st.container(border=True):
                    col1, col2, col3 = st.columns([4, 2, 2])
                    with col1:
                        st.subheader(f"📚 {nombre}")
                    with col2:
                        st.write(f"⏱️ {horas} horas")
                    with col3:
                        if st.button("✏️ Editar", key=f"editar_curso_{curso_id}"):
                            st.session_state.curso_editando = curso_id
                            st.rerun()
                    if st.session_state.curso_editando == curso_id:
                        with st.form(f"form_editar_curso_{curso_id}"):
                            nuevo_nombre = st.text_input("Nombre del curso", value=nombre)
                            horas_opciones = list(range(40, 331))
                            indice = horas_opciones.index(horas) if horas in horas_opciones else 0
                            nuevas_horas = st.selectbox("Número de horas", horas_opciones, index=indice, format_func=lambda x: f"{x} horas")
                            guardar = st.form_submit_button("💾 Guardar cambios")
                            if guardar:
                                if nuevo_nombre.strip() == "":
                                    st.error("El nombre no puede estar vacío.")
                                else:
                                    try:
                                        cursor.execute("UPDATE productos SET nombre = ?, horas = ? WHERE id = ?", (nuevo_nombre.strip(), nuevas_horas, curso_id))
                                        conexion.commit()
                                        st.session_state.curso_editando = None
                                        st.success("✅ Curso actualizado correctamente.")
                                        st.rerun()
                                    except IntegrityError:
                                        conexion.rollback()
                                        st.error("⚠️ Ese curso ya existe.")
                    if st.button("🗑️ Eliminar", key=f"eliminar_curso_{curso_id}"):
                        st.session_state.curso_eliminando = curso_id
                        st.rerun()
                    if st.session_state.curso_eliminando == curso_id:
                        cantidad = cantidad_servicios_curso(curso_id)
                        if cantidad > 0:
                            st.error("🚫 No se puede eliminar este curso porque está asociado a trabajos.")
                        else:
                            confirmar = st.checkbox("Confirmo que deseo eliminar este curso.", key=f"confirmar_curso_{curso_id}")
                            if confirmar:
                                if st.button("🗑️ Confirmar eliminación", key=f"confirmar_eliminar_curso_{curso_id}"):
                                    cursor.execute("DELETE FROM productos WHERE id = ?", (curso_id,))
                                    conexion.commit()
                                    st.session_state.curso_eliminando = None
                                    st.success("✅ Curso eliminado correctamente.")
                                    st.rerun()


# ============================================================
# NUEVO TRABAJO
# ============================================================

elif opcion == "🆕 Nuevo trabajo":
    st.header("🆕 Nuevo trabajo")
    st.write("Registra aquí un curso normal o una PROMO y define su plan de pagos.")
    cursor.execute("SELECT id, cedula, nombres FROM docentes ORDER BY nombres")
    docentes = cursor.fetchall()
    cursor.execute("SELECT id, nombre, horas FROM productos ORDER BY nombre")
    cursos = cursor.fetchall()

    if not docentes:
        st.warning("⚠️ Primero debes registrar al menos un docente.")
    elif not cursos:
        st.warning("⚠️ Primero debes registrar al menos un curso.")
    else:
        opciones_docentes = {f"{docente[2]} - Cédula: {docente[1]}": docente[0] for docente in docentes}
        opciones_cursos = {curso[1]: curso[0] for curso in cursos}
        opciones_trabajo = ["PROMO - 330 horas"]
        opciones_trabajo.extend([curso[1] for curso in cursos])

        with st.form("form_nuevo_trabajo"):
            docente_seleccionado = st.selectbox("👨‍🏫 Docente", list(opciones_docentes.keys()))
            docente_id = opciones_docentes[docente_seleccionado]
            trabajo_seleccionado = st.selectbox("📚 Curso / Trabajo", opciones_trabajo)
            es_promo = trabajo_seleccionado == "PROMO - 330 horas"

            promo_cantidad = 0
            promo_curso_1 = None
            promo_curso_2 = None
            promo_curso_3 = None
            promo_curso_4 = None
            promo_curso_5 = None
            producto_id = None
            cursos_ids = []

            if es_promo:
                st.success("🎓 Has seleccionado la PROMO ESPECIAL de 330 horas.")
                promo_cantidad = st.selectbox("🎓 ¿Cuántos cursos tendrá la PROMO?", [3, 5], format_func=lambda x: f"PROMO de {x} cursos")
                st.info("Selecciona los cursos que formarán parte de esta promoción.")
                promo_curso_1_nombre = st.selectbox("📘 Curso 1", list(opciones_cursos.keys()), key="promo_curso_1")
                promo_curso_1 = opciones_cursos[promo_curso_1_nombre]
                promo_curso_2_nombre = st.selectbox("📗 Curso 2", list(opciones_cursos.keys()), key="promo_curso_2")
                promo_curso_2 = opciones_cursos[promo_curso_2_nombre]
                promo_curso_3_nombre = st.selectbox("📙 Curso 3", list(opciones_cursos.keys()), key="promo_curso_3")
                promo_curso_3 = opciones_cursos[promo_curso_3_nombre]

                if promo_cantidad == 5:
                    promo_curso_4_nombre = st.selectbox("📕 Curso 4", list(opciones_cursos.keys()), key="promo_curso_4")
                    promo_curso_4 = opciones_cursos[promo_curso_4_nombre]
                    promo_curso_5_nombre = st.selectbox("📔 Curso 5", list(opciones_cursos.keys()), key="promo_curso_5")
                    promo_curso_5 = opciones_cursos[promo_curso_5_nombre]

                lista_promo = [promo_curso_1, promo_curso_2, promo_curso_3]
                if promo_cantidad == 5:
                    lista_promo.extend([promo_curso_4, promo_curso_5])

                cursos_ids = lista_promo

                if cursos_repetidos(lista_promo):
                    st.error("⚠️ No puedes seleccionar el mismo curso más de una vez dentro de la misma PROMO.")

                cursos_ya_registrados = []
                for curso_id in lista_promo:
                    if docente_tiene_curso(docente_id, curso_id):
                        cursor.execute("SELECT nombre FROM productos WHERE id = ?", (curso_id,))
                        resultado = cursor.fetchone()
                        if resultado:
                            cursos_ya_registrados.append(resultado[0])
                if cursos_ya_registrados:
                    st.error("⚠️ Este docente ya tiene registrado(s) estos curso(s):")
                    for nombre in cursos_ya_registrados:
                        st.write(f"• {nombre}")
                    st.warning("No se puede registrar una PROMO que contenga cursos que el docente ya tenga matriculados.")
            else:
                producto_id = opciones_cursos[trabajo_seleccionado]
                curso_actual = next((curso for curso in cursos if curso[0] == producto_id), None)
                if curso_actual:
                    st.info(f"📚 **Curso:** {curso_actual[1]}  \n⏱️ **Horas:** {curso_actual[2]} horas")
                if docente_tiene_curso(docente_id, producto_id):
                    st.error("⚠️ Este docente ya está matriculado en este curso.")
                    st.warning("No puedes registrar nuevamente este mismo curso para este docente.")
                cursos_ids = [producto_id]

            st.divider()
            fecha_inicio_trabajo = st.date_input("📅 Fecha de inicio del trabajo", value=date.today())
            precio = st.number_input("💰 Precio total", min_value=0.0, step=0.01, format="%.2f")
            
            st.divider()
            st.subheader("💳 Plan de pagos")
            numero_pagos = st.selectbox("Seleccione el plan de pagos", [1, 2, 3], format_func=lambda x: "1 pago" if x == 1 else f"{x} pagos")
            
            st.info("📅 **Las fechas estimadas son obligatorias** para cada cuota del plan seleccionado.")
            fecha_est_1 = st.date_input("📅 Fecha estimada pago 1", value=None)
            fecha_est_2 = None
            fecha_est_3 = None
            if numero_pagos >= 2:
                fecha_est_2 = st.date_input("📅 Fecha estimada pago 2", value=None)
            if numero_pagos == 3:
                fecha_est_3 = st.date_input("📅 Fecha estimada pago 3", value=None)

            st.info("💡 Podrás registrar los montos y fechas reales de cada pago en el módulo **Trabajos en proceso** conforme el docente vaya cancelando.")

            st.divider()
            observaciones = st.text_area("📝 Observaciones")
            guardar = st.form_submit_button("💾 Guardar nuevo trabajo")

            if guardar:
                errores = []
                if precio <= 0:
                    errores.append("El precio debe ser mayor que 0.")
                if fecha_est_1 is None:
                    errores.append("Debes ingresar la fecha estimada del pago 1.")
                if numero_pagos >= 2 and fecha_est_2 is None:
                    errores.append("Debes ingresar la fecha estimada del pago 2.")
                if numero_pagos == 3 and fecha_est_3 is None:
                    errores.append("Debes ingresar la fecha estimada del pago 3.")
                if fecha_est_1 and fecha_est_2 and fecha_est_2 < fecha_est_1:
                    errores.append("La fecha estimada del pago 2 no puede ser anterior a la del pago 1.")
                if fecha_est_2 and fecha_est_3 and fecha_est_3 < fecha_est_2:
                    errores.append("La fecha estimada del pago 3 no puede ser anterior a la del pago 2.")
                if fecha_est_1 and fecha_est_3 and fecha_est_3 < fecha_est_1:
                    errores.append("La fecha estimada del pago 3 no puede ser anterior a la del pago 1.")

                if es_promo:
                    lista_promo = [promo_curso_1, promo_curso_2, promo_curso_3]
                    if promo_cantidad == 5:
                        lista_promo.extend([promo_curso_4, promo_curso_5])
                    if cursos_repetidos(lista_promo):
                        errores.append("No puedes repetir un curso dentro de la misma PROMO.")
                    cursos_duplicados = []
                    for curso_id in lista_promo:
                        if docente_tiene_curso(docente_id, curso_id):
                            cursor.execute("SELECT nombre FROM productos WHERE id = ?", (curso_id,))
                            resultado = cursor.fetchone()
                            if resultado:
                                cursos_duplicados.append(resultado[0])
                    if cursos_duplicados:
                        errores.append("El docente ya tiene registrado uno o más cursos seleccionados en esta PROMO.")
                else:
                    if docente_tiene_curso(docente_id, producto_id):
                        errores.append("Este docente ya está matriculado en este curso. No puede registrarse nuevamente.")
                
                if errores:
                    for error in errores:
                        st.error(f"⚠️ {error}")
                else:
                    if es_promo:
                        tipo_trabajo = "PROMO"
                        descripcion = f"PROMO de {promo_cantidad} cursos"
                        producto_id_guardar = None
                    else:
                        tipo_trabajo = "CURSO"
                        descripcion = trabajo_seleccionado
                        producto_id_guardar = producto_id
                    
                    fecha_inicio_texto = fecha_inicio_trabajo.isoformat()
                    f_est_1_txt = fecha_est_1.isoformat()
                    f_est_2_txt = fecha_est_2.isoformat() if fecha_est_2 else None
                    f_est_3_txt = fecha_est_3.isoformat() if fecha_est_3 else None

                    try:
                        cursor.execute(
                            """
                            INSERT INTO servicios (
                                docente_id, descripcion, precio, abonos, observaciones, producto_id, 
                                numero_pagos, fecha_pago_1, fecha_pago_2, fecha_pago_3, 
                                monto_pago_1, monto_pago_2, monto_pago_3, tipo_trabajo, 
                                promo_cantidad, promo_curso_1, promo_curso_2, promo_curso_3, 
                                promo_curso_4, promo_curso_5, fecha_inicio_trabajo,
                                fecha_estimada_1, fecha_estimada_2, fecha_estimada_3
                            ) VALUES (?, ?, ?, 0.0, ?, ?, ?, NULL, NULL, NULL, 0.0, 0.0, 0.0, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            RETURNING id
                            """,
                            (
                                docente_id, descripcion, precio, observaciones.strip(), producto_id_guardar, 
                                numero_pagos, tipo_trabajo, 
                                promo_cantidad, promo_curso_1, promo_curso_2, promo_curso_3, 
                                promo_curso_4, promo_curso_5, fecha_inicio_texto,
                                f_est_1_txt, f_est_2_txt, f_est_3_txt
                            )
                        )
                        servicio_id = cursor.fetchone()[0]
                        if cursos_ids:
                            inicializar_estados(servicio_id, cursos_ids)
                        conexion.commit()
                        st.success("✅ Nuevo trabajo registrado correctamente.")
                        if es_promo:
                            st.info(f"🎓 PROMO de {promo_cantidad} cursos registrada correctamente.")
                        st.success(f"💵 Saldo pendiente inicial: ${precio:.2f}")
                    except IntegrityError as e:
                        conexion.rollback()
                        st.error("❌ No se pudo guardar el trabajo.")
                        st.caption(f"Detalle: {e}")


# ============================================================
# TRABAJOS EN PROCESO (con exportación individual y general)
# ============================================================

elif opcion == "📋 Trabajos en proceso":

    st.header("📋 Trabajos en proceso")

    # ========== FUNCIONES DE EXPORTACIÓN ==========

    def generar_reporte_excel_individual(docente_id, lista_servicios):
        """Genera un Excel con los trabajos de un solo docente."""
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ---------- HOJA DETALLE ----------
        ws_detalle = wb.active
        ws_detalle.title = "Detalle por curso"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F5597")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')

        headers = [
            "Docente", "Cédula", "Teléfono", "Usuario",
            "ID Trabajo", "Tipo", "Curso", "Horas",
            "Estado", "Precio Total", "Abonos", "Saldo",
            "Fecha Inicio", "Observaciones"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws_detalle.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        row_num = 2
        for servicio in lista_servicios:
            servicio_id = servicio[0]
            cedula = servicio[1] or ""
            nombres = servicio[2] or ""
            telefono = servicio[3] or ""
            usuario = servicio[4] or ""
            precio = servicio[9] or 0.0
            abonos = servicio[10] or 0.0
            observaciones = servicio[11] or ""
            tipo_trabajo = servicio[19] or "CURSO"
            fecha_inicio = servicio[26] or ""

            cursos_info = obtener_cursos_servicio(servicio_id)
            estados_dict = obtener_estados_servicio(servicio_id)

            if cursos_info:
                for curso_id, curso_nombre, curso_horas in cursos_info:
                    estado = estados_dict.get(curso_id, "Matriculado")
                    saldo = precio - abonos
                    row_data = [
                        nombres, cedula, telefono, usuario,
                        servicio_id,
                        "PROMO" if tipo_trabajo == "PROMO" else "CURSO",
                        curso_nombre, curso_horas,
                        estado,
                        precio, abonos, saldo,
                        fecha_inicio, observaciones
                    ]
                    for col, value in enumerate(row_data, 1):
                        cell = ws_detalle.cell(row=row_num, column=col, value=value)
                        cell.border = thin_border
                        cell.alignment = left_alignment if col in [1, 7] else center_alignment
                    row_num += 1
            else:
                saldo = precio - abonos
                row_data = [
                    nombres, cedula, telefono, usuario,
                    servicio_id,
                    "PROMO" if tipo_trabajo == "PROMO" else "CURSO",
                    servicio[7] or "Sin curso", 0,
                    "Sin estado",
                    precio, abonos, saldo,
                    fecha_inicio, observaciones
                ]
                for col, value in enumerate(row_data, 1):
                    cell = ws_detalle.cell(row=row_num, column=col, value=value)
                    cell.border = thin_border
                    cell.alignment = left_alignment if col in [1, 7] else center_alignment
                row_num += 1

        # Ajustar anchos
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            if col in [1, 3, 7]:
                ws_detalle.column_dimensions[col_letter].width = 25
            elif col in [2, 4, 5, 6, 10, 11, 12]:
                ws_detalle.column_dimensions[col_letter].width = 15
            elif col == 8:
                ws_detalle.column_dimensions[col_letter].width = 10
            elif col == 9:
                ws_detalle.column_dimensions[col_letter].width = 18
            elif col == 13:
                ws_detalle.column_dimensions[col_letter].width = 15
            elif col == 14:
                ws_detalle.column_dimensions[col_letter].width = 35
            else:
                ws_detalle.column_dimensions[col_letter].width = 15
        ws_detalle.freeze_panes = 'A2'

        # ---------- HOJA RESUMEN ----------
        ws_resumen = wb.create_sheet("Resumen del docente")
        headers_resumen = [
            "Docente", "Cédula", "Teléfono", "Usuario",
            "Total Trabajos", "Total Cursos",
            "Suma Precios", "Suma Abonos", "Saldo Total"
        ]
        for col, header in enumerate(headers_resumen, 1):
            cell = ws_resumen.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        total_trabajos = len(lista_servicios)
        total_cursos = 0
        suma_precios = sum(s[9] for s in lista_servicios)
        suma_abonos = sum(s[10] for s in lista_servicios)
        saldo_total = suma_precios - suma_abonos
        for servicio in lista_servicios:
            servicio_id = servicio[0]
            estados_dict = obtener_estados_servicio(servicio_id)
            total_cursos += len(estados_dict)

        row_data = [
            nombres, cedula, telefono, usuario,
            total_trabajos, total_cursos,
            suma_precios, suma_abonos, saldo_total
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws_resumen.cell(row=2, column=col, value=value)
            cell.border = thin_border
            cell.alignment = center_alignment
        # Ajustar anchos
        for col in range(1, len(headers_resumen) + 1):
            col_letter = get_column_letter(col)
            if col in [1, 4]:
                ws_resumen.column_dimensions[col_letter].width = 25
            elif col in [2, 3]:
                ws_resumen.column_dimensions[col_letter].width = 18
            else:
                ws_resumen.column_dimensions[col_letter].width = 15
        ws_resumen.freeze_panes = 'A2'

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generar_reporte_excel_general(servicios):
        """Genera el reporte general (todos los docentes)"""
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from collections import defaultdict

        wb = Workbook()

        # ------------------- HOJA DETALLE -------------------
        ws_detalle = wb.active
        ws_detalle.title = "Detalle por curso"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2F5597")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')

        headers = [
            "Docente", "Cédula", "Teléfono", "Usuario",
            "ID Trabajo", "Tipo", "Curso", "Horas",
            "Estado", "Precio Total", "Abonos", "Saldo",
            "Fecha Inicio", "Observaciones"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws_detalle.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        grupos = defaultdict(list)
        for s in servicios:
            docente_id = s[28]
            grupos[docente_id].append(s)

        row_num = 2
        for docente_id, lista_servicios in grupos.items():
            for servicio in lista_servicios:
                servicio_id = servicio[0]
                cedula = servicio[1] or ""
                nombres = servicio[2] or ""
                telefono = servicio[3] or ""
                usuario = servicio[4] or ""
                precio = servicio[9] or 0.0
                abonos = servicio[10] or 0.0
                observaciones = servicio[11] or ""
                tipo_trabajo = servicio[19] or "CURSO"
                fecha_inicio = servicio[26] or ""

                cursos_info = obtener_cursos_servicio(servicio_id)
                estados_dict = obtener_estados_servicio(servicio_id)

                if cursos_info:
                    for curso_id, curso_nombre, curso_horas in cursos_info:
                        estado = estados_dict.get(curso_id, "Matriculado")
                        saldo = precio - abonos
                        row_data = [
                            nombres, cedula, telefono, usuario,
                            servicio_id,
                            "PROMO" if tipo_trabajo == "PROMO" else "CURSO",
                            curso_nombre, curso_horas,
                            estado,
                            precio, abonos, saldo,
                            fecha_inicio, observaciones
                        ]
                        for col, value in enumerate(row_data, 1):
                            cell = ws_detalle.cell(row=row_num, column=col, value=value)
                            cell.border = thin_border
                            cell.alignment = left_alignment if col in [1, 7] else center_alignment
                        row_num += 1
                else:
                    saldo = precio - abonos
                    row_data = [
                        nombres, cedula, telefono, usuario,
                        servicio_id,
                        "PROMO" if tipo_trabajo == "PROMO" else "CURSO",
                        servicio[7] or "Sin curso", 0,
                        "Sin estado",
                        precio, abonos, saldo,
                        fecha_inicio, observaciones
                    ]
                    for col, value in enumerate(row_data, 1):
                        cell = ws_detalle.cell(row=row_num, column=col, value=value)
                        cell.border = thin_border
                        cell.alignment = left_alignment if col in [1, 7] else center_alignment
                    row_num += 1

        # Ajustar anchos
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            if col in [1, 3, 7]:
                ws_detalle.column_dimensions[col_letter].width = 25
            elif col in [2, 4, 5, 6, 10, 11, 12]:
                ws_detalle.column_dimensions[col_letter].width = 15
            elif col == 8:
                ws_detalle.column_dimensions[col_letter].width = 10
            elif col == 9:
                ws_detalle.column_dimensions[col_letter].width = 18
            elif col == 13:
                ws_detalle.column_dimensions[col_letter].width = 15
            elif col == 14:
                ws_detalle.column_dimensions[col_letter].width = 35
            else:
                ws_detalle.column_dimensions[col_letter].width = 15
        ws_detalle.freeze_panes = 'A2'

        # ------------------- HOJA RESUMEN -------------------
        ws_resumen = wb.create_sheet("Resumen por docente")
        headers_resumen = [
            "Docente", "Cédula", "Teléfono", "Usuario",
            "Total Trabajos", "Total Cursos",
            "Suma Precios", "Suma Abonos", "Saldo Total",
            "¿Terminado?", "¿En Proceso?"
        ]
        for col, header in enumerate(headers_resumen, 1):
            cell = ws_resumen.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        row_num = 2
        for docente_id, lista_servicios in grupos.items():
            primer = lista_servicios[0]
            nombres = primer[2] or ""
            cedula = primer[1] or ""
            telefono = primer[3] or ""
            usuario = primer[4] or ""

            total_trabajos = len(lista_servicios)
            total_cursos = 0
            suma_precios = sum(s[9] for s in lista_servicios)
            suma_abonos = sum(s[10] for s in lista_servicios)
            saldo_total = suma_precios - suma_abonos

            terminado = True
            en_proceso = False
            for servicio in lista_servicios:
                servicio_id = servicio[0]
                estados_dict = obtener_estados_servicio(servicio_id)
                for estado in estados_dict.values():
                    total_cursos += 1
                    if estado not in ["Entregado", "Terminado"]:
                        terminado = False
                    if estado in ["En proceso", "Matriculado"]:
                        en_proceso = True
            if total_cursos == 0:
                terminado = False
            if terminado and saldo_total != 0:
                terminado = False

            row_data = [
                nombres, cedula, telefono, usuario,
                total_trabajos, total_cursos,
                suma_precios, suma_abonos, saldo_total,
                "✅" if terminado else "❌",
                "✅" if en_proceso else "❌"
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws_resumen.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                cell.alignment = left_alignment if col == 1 else center_alignment
            row_num += 1

        # Ajustar anchos resumen
        for col in range(1, len(headers_resumen) + 1):
            col_letter = get_column_letter(col)
            if col in [1, 4]:
                ws_resumen.column_dimensions[col_letter].width = 25
            elif col in [2, 10, 11]:
                ws_resumen.column_dimensions[col_letter].width = 15
            elif col in [3, 5, 6]:
                ws_resumen.column_dimensions[col_letter].width = 18
            else:
                ws_resumen.column_dimensions[col_letter].width = 15
        ws_resumen.freeze_panes = 'A2'

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ========== OBTENER DATOS ==========

    cursor.execute("SELECT id, nombre FROM productos ORDER BY nombre")
    todos_los_cursos = cursor.fetchall()
    opciones_cursos = {c[1]: c[0] for c in todos_los_cursos}
    nombres_cursos_lista = list(opciones_cursos.keys())

    cursor.execute(
        """
        SELECT
            servicios.id,
            docentes.cedula,
            docentes.nombres,
            docentes.telefono,
            docentes.usuario,
            docentes.clave,
            docentes.correo,
            productos.nombre,
            productos.horas,
            servicios.precio,
            servicios.abonos,
            servicios.observaciones,
            servicios.numero_pagos,
            servicios.fecha_pago_1,
            servicios.fecha_pago_2,
            servicios.fecha_pago_3,
            servicios.monto_pago_1,
            servicios.monto_pago_2,
            servicios.monto_pago_3,
            servicios.tipo_trabajo,
            servicios.promo_cantidad,
            servicios.promo_curso_1,
            servicios.promo_curso_2,
            servicios.promo_curso_3,
            servicios.promo_curso_4,
            servicios.promo_curso_5,
            servicios.fecha_inicio_trabajo,
            servicios.producto_id,
            servicios.docente_id,
            servicios.fecha_estimada_1,
            servicios.fecha_estimada_2,
            servicios.fecha_estimada_3
        FROM servicios
        LEFT JOIN docentes ON servicios.docente_id = docentes.id
        LEFT JOIN productos ON servicios.producto_id = productos.id
        ORDER BY docentes.nombres, servicios.id DESC
        """
    )
    servicios = cursor.fetchall()

    if not servicios:
        st.info("Todavía no hay trabajos registrados.")
    else:
        # Agrupar por docente
        grupos = defaultdict(list)
        for s in servicios:
            docente_id = s[28]
            grupos[docente_id].append(s)

        # ----- Cálculo de métricas -----
        docentes_terminados = 0
        docentes_en_proceso = 0

        for docente_id, lista_servicios in grupos.items():
            saldo_total = sum(s[9] - s[10] for s in lista_servicios)
            todos_terminados = True
            tiene_en_proceso = False

            for servicio in lista_servicios:
                servicio_id = servicio[0]
                estados_dict = obtener_estados_servicio(servicio_id)
                for estado in estados_dict.values():
                    if estado not in ["Entregado", "Terminado"]:
                        todos_terminados = False
                    if estado in ["En proceso", "Matriculado"]:
                        tiene_en_proceso = True

            if todos_terminados and saldo_total == 0:
                docentes_terminados += 1
            if tiene_en_proceso:
                docentes_en_proceso += 1

        # Mostrar estadísticas y botón de exportación general
        col_est1, col_est2, col_exp = st.columns([2, 2, 1])
        with col_est1:
            st.write(f"**📌 Docentes con trabajos terminados:** {docentes_terminados}")
        with col_est2:
            st.write(f"**📌 Docentes con trabajos en proceso:** {docentes_en_proceso}")
        with col_exp:
            if st.button("📥 Exportar todos", use_container_width=True, type="primary"):
                excel_data = generar_reporte_excel_general(servicios)
                st.download_button(
                    label="📥 Descargar reporte general",
                    data=excel_data,
                    file_name=f"reporte_trabajos_{date.today().isoformat()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        st.divider()

        # Botones para abrir/colapsar todos
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("📂 Abrir todos los trabajos", use_container_width=True):
                st.session_state.trabajos_abiertos = set(grupos.keys())
                st.rerun()
        with col_btn2:
            if st.button("📁 Colapsar todos los trabajos", use_container_width=True):
                st.session_state.trabajos_abiertos = set()
                st.rerun()

        st.divider()

        # Iterar por cada docente
        for docente_id, lista_servicios in grupos.items():
            primer = lista_servicios[0]
            cedula = primer[1]
            nombres = primer[2]
            telefono = primer[3]
            usuario = primer[4]
            clave = primer[5]

            total_trabajos = len(lista_servicios)
            suma_saldos = sum(s[9] - s[10] for s in lista_servicios)

            abierto = docente_id in st.session_state.trabajos_abiertos

            # Encabezado del docente con botón de exportación individual
            col1, col2, col3 = st.columns([4, 3, 1])
            with col1:
                st.subheader(f"👨‍🏫 {nombres or 'Sin docente'}")
                st.caption(f"Cédula: {cedula or 'N/A'}  |  Teléfono: {telefono or 'N/A'}")
                st.write(f"**Usuario:** {usuario or 'Sin usuario'}  |  **Clave:** {clave or 'Sin clave'}")
            with col2:
                st.write(f"**Total de trabajos:** {total_trabajos}")
                st.write(f"**Suma total de saldos pendientes:** ${suma_saldos:.2f}")
            with col3:
                st.download_button(
                    label="📥 Exportar",
                    data=generar_reporte_excel_individual(docente_id, lista_servicios),
                    file_name=f"reporte_{nombres.replace(' ', '_')}_{date.today().isoformat()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"export_{docente_id}",
                    use_container_width=True
                )

            # Botón Mostrar/Ocultar
            col_mostrar = st.columns([1])[0]
            with col_mostrar:
                texto_boton = "📁 Ocultar" if abierto else "📂 Mostrar"
                if st.button(texto_boton, key=f"mostrar_trabajos_{docente_id}", use_container_width=True):
                    if abierto:
                        st.session_state.trabajos_abiertos.remove(docente_id)
                    else:
                        st.session_state.trabajos_abiertos.add(docente_id)
                    st.rerun()

            if abierto:
                st.divider()
                for servicio in lista_servicios:
                    servicio_id = servicio[0]
                    nombre_curso = servicio[7]
                    horas = servicio[8]
                    precio = servicio[9]
                    abonos = servicio[10]
                    observaciones = servicio[11]
                    numero_pagos = servicio[12]
                    fecha_pago_1 = servicio[13]
                    fecha_pago_2 = servicio[14]
                    fecha_pago_3 = servicio[15]
                    monto_pago_1 = servicio[16]
                    monto_pago_2 = servicio[17]
                    monto_pago_3 = servicio[18]
                    tipo_trabajo = servicio[19]
                    promo_cantidad = servicio[20]
                    promo_curso_1 = servicio[21]
                    promo_curso_2 = servicio[22]
                    promo_curso_3 = servicio[23]
                    promo_curso_4 = servicio[24]
                    promo_curso_5 = servicio[25]
                    fecha_inicio = servicio[26]
                    producto_id_actual = servicio[27]
                    fecha_estimada_1 = servicio[29]
                    fecha_estimada_2 = servicio[30]
                    fecha_estimada_3 = servicio[31]

                    saldo = precio - abonos

                    cursos_info = obtener_cursos_servicio(servicio_id)
                    estados_dict = obtener_estados_servicio(servicio_id)

                    with st.container(border=True):
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.write("**📚 Trabajo / Curso**")
                            if tipo_trabajo == "PROMO":
                                st.write(f"🎓 PROMO de {promo_cantidad} cursos")
                                st.write(f"⏱️ 330 horas")
                            else:
                                st.write(nombre_curso or "Sin curso")
                                if horas:
                                    st.write(f"⏱️ {horas} horas")
                            if fecha_inicio:
                                st.write(f"📅 **Inicio:** {fecha_inicio}")
                        with col2:
                            st.write("**💰 Precio total**")
                            st.write(f"${precio:.2f}")
                            st.write("**💵 Abonos Totales**")
                            st.write(f"${abonos:.2f}")
                        with col3:
                            st.write("**💳 Saldo pendiente**")
                            if saldo > 0:
                                st.warning(f"${saldo:.2f}")
                            else:
                                st.success("$0.00")

                        if cursos_info:
                            st.write("**📌 Cursos y estados:**")
                            for curso_id, curso_nombre, _ in cursos_info:
                                estado_actual = estados_dict.get(curso_id, "Matriculado")
                                color = {
                                    "Matriculado": "blue",
                                    "En proceso": "orange",
                                    "Terminado": "green",
                                    "Entregado": "green",
                                    "No salió": "red",
                                    "Cancelado": "red",
                                    "Próxima matrícula": "purple"
                                }.get(estado_actual, "gray")
                                st.markdown(
                                    f"• {curso_nombre}: <span style='color:{color}; font-weight:bold;'>{estado_actual}</span>",
                                    unsafe_allow_html=True
                                )

                        if tipo_trabajo == "PROMO" and not cursos_info:
                            cursos_promo_ids = [promo_curso_1, promo_curso_2, promo_curso_3, promo_curso_4, promo_curso_5]
                            nombres_promo = []
                            for cid in cursos_promo_ids:
                                if cid is not None:
                                    cursor.execute("SELECT nombre, horas FROM productos WHERE id = ?", (cid,))
                                    resultado = cursor.fetchone()
                                    if resultado:
                                        nombres_promo.append(resultado)
                            if nombres_promo:
                                st.write("**📚 Cursos incluidos:**")
                                for idx, (nom, hrs) in enumerate(nombres_promo, start=1):
                                    st.caption(f"{idx}. {nom} — {hrs} horas")

                        st.write("---")
                        st.write(f"**🗓️ Detalle del plan de {numero_pagos} pago(s):**")
                        col_f1, col_f2, col_f3 = st.columns(3)
                        with col_f1:
                            if monto_pago_1 > 0:
                                if fecha_pago_1:
                                    st.write(f"**Pago 1:** ${monto_pago_1:.2f}  \n📅 {fecha_pago_1}")
                                else:
                                    st.write(f"**Pago 1:** ${monto_pago_1:.2f}  \n📅 *Sin fecha*")
                            else:
                                if fecha_estimada_1:
                                    st.write(f"**Pago 1:** *Pendiente* (estimado: {fecha_estimada_1})")
                                else:
                                    st.write("**Pago 1:** *Pendiente*")
                        with col_f2:
                            if numero_pagos >= 2:
                                if monto_pago_2 > 0:
                                    if fecha_pago_2:
                                        st.write(f"**Pago 2:** ${monto_pago_2:.2f}  \n📅 {fecha_pago_2}")
                                    else:
                                        st.write(f"**Pago 2:** ${monto_pago_2:.2f}  \n📅 *Sin fecha*")
                                else:
                                    if fecha_estimada_2:
                                        st.write(f"**Pago 2:** *Pendiente* (estimado: {fecha_estimada_2})")
                                    else:
                                        st.write("**Pago 2:** *Pendiente*")
                        with col_f3:
                            if numero_pagos == 3:
                                if monto_pago_3 > 0:
                                    if fecha_pago_3:
                                        st.write(f"**Pago 3:** ${monto_pago_3:.2f}  \n📅 {fecha_pago_3}")
                                    else:
                                        st.write(f"**Pago 3:** ${monto_pago_3:.2f}  \n📅 *Sin fecha*")
                                else:
                                    if fecha_estimada_3:
                                        st.write(f"**Pago 3:** *Pendiente* (estimado: {fecha_estimada_3})")
                                    else:
                                        st.write("**Pago 3:** *Pendiente*")

                        if observaciones:
                            st.info(f"**📝 Observaciones:** {observaciones}")

                        col_btn1, col_btn2 = st.columns(2)
                        with col_btn1:
                            if st.button("✏️ Editar / Registrar Pagos", key=f"btn_edit_{servicio_id}", use_container_width=True):
                                st.session_state.servicio_editando = servicio_id
                                st.session_state.servicio_eliminando = None
                                st.rerun()
                        with col_btn2:
                            if st.button("🗑️ Eliminar", key=f"btn_del_{servicio_id}", use_container_width=True):
                                st.session_state.servicio_eliminando = servicio_id
                                st.session_state.servicio_editando = None
                                st.rerun()

                        if st.session_state.servicio_editando == servicio_id:
                            st.info("👇 Registra los pagos, fechas, edita el precio o cambia los estados de los cursos a continuación:")
                            with st.form(f"form_edit_srv_{servicio_id}"):
                                nuevo_producto_id = producto_id_actual
                                if tipo_trabajo != "PROMO" and nombres_cursos_lista:
                                    indice_actual = 0
                                    if nombre_curso in nombres_cursos_lista:
                                        indice_actual = nombres_cursos_lista.index(nombre_curso)
                                    curso_elegido = st.selectbox("📚 Corregir Curso", nombres_cursos_lista, index=indice_actual)
                                    nuevo_producto_id = opciones_cursos[curso_elegido]

                                nuevo_precio = st.number_input("💰 Precio total", min_value=0.0, value=float(precio), step=0.01)

                                st.subheader("📌 Estados de los cursos")
                                cursos_actuales = obtener_cursos_servicio(servicio_id)
                                estados_actuales = obtener_estados_servicio(servicio_id)
                                nuevos_estados = {}
                                for curso_id, curso_nombre, _ in cursos_actuales:
                                    estado_actual = estados_actuales.get(curso_id, "Matriculado")
                                    nuevo_estado = st.selectbox(
                                        f"Estado de {curso_nombre}",
                                        ESTADOS_POSIBLES,
                                        index=ESTADOS_POSIBLES.index(estado_actual) if estado_actual in ESTADOS_POSIBLES else 0,
                                        key=f"estado_{servicio_id}_{curso_id}"
                                    )
                                    nuevos_estados[curso_id] = nuevo_estado

                                st.divider()
                                st.subheader(f"💳 Registro de Abonos ({numero_pagos} pago(s) planificados)")

                                val_f1 = date.fromisoformat(fecha_pago_1) if fecha_pago_1 else None
                                val_f2 = date.fromisoformat(fecha_pago_2) if fecha_pago_2 else None
                                val_f3 = date.fromisoformat(fecha_pago_3) if fecha_pago_3 else None

                                st.write("**Detalle del Pago 1:**")
                                col_ep1_1, col_ep1_2 = st.columns(2)
                                with col_ep1_1:
                                    nuevo_monto_p1 = st.number_input("💵 Monto pago 1", min_value=0.0, value=float(monto_pago_1), step=0.01, key=f"edit_m_p1_{servicio_id}")
                                with col_ep1_2:
                                    nueva_fecha_p1 = st.date_input("📅 Fecha pago 1 (opcional)", value=val_f1, key=f"edit_f_p1_{servicio_id}")

                                nuevo_monto_p2 = 0.0
                                nueva_fecha_p2 = None
                                nuevo_monto_p3 = 0.0
                                nueva_fecha_p3 = None

                                if numero_pagos >= 2:
                                    st.write("**Detalle del Pago 2:**")
                                    col_ep2_1, col_ep2_2 = st.columns(2)
                                    with col_ep2_1:
                                        nuevo_monto_p2 = st.number_input("💵 Monto pago 2", min_value=0.0, value=float(monto_pago_2), step=0.01, key=f"edit_m_p2_{servicio_id}")
                                    with col_ep2_2:
                                        nueva_fecha_p2 = st.date_input("📅 Fecha pago 2 (opcional)", value=val_f2, key=f"edit_f_p2_{servicio_id}")

                                if numero_pagos == 3:
                                    st.write("**Detalle del Pago 3:**")
                                    col_ep3_1, col_ep3_2 = st.columns(2)
                                    with col_ep3_1:
                                        nuevo_monto_p3 = st.number_input("💵 Monto pago 3", min_value=0.0, value=float(monto_pago_3), step=0.01, key=f"edit_m_p3_{servicio_id}")
                                    with col_ep3_2:
                                        nueva_fecha_p3 = st.date_input("📅 Fecha pago 3 (opcional)", value=val_f3, key=f"edit_f_p3_{servicio_id}")

                                nuevos_abonos = nuevo_monto_p1 + nuevo_monto_p2 + nuevo_monto_p3
                                nuevas_obs = st.text_area("📝 Observaciones", value=observaciones if observaciones else "")

                                guardar_edicion = st.form_submit_button("💾 Guardar cambios")

                                if guardar_edicion:
                                    errores_edit = []
                                    if nuevos_abonos > nuevo_precio:
                                        errores_edit.append("La suma de los abonos no puede ser mayor que el precio total.")
                                    if nueva_fecha_p1 and nueva_fecha_p2 and nueva_fecha_p2 < nueva_fecha_p1:
                                        errores_edit.append("La fecha del pago 2 no puede ser anterior a la fecha del pago 1.")
                                    if nueva_fecha_p2 and nueva_fecha_p3 and nueva_fecha_p3 < nueva_fecha_p2:
                                        errores_edit.append("La fecha del pago 3 no puede ser anterior a la fecha del pago 2.")

                                    if errores_edit:
                                        for err in errores_edit:
                                            st.error(f"⚠️ {err}")
                                    else:
                                        f1_txt = nueva_fecha_p1.isoformat() if (nuevo_monto_p1 > 0 and nueva_fecha_p1) else None
                                        f2_txt = nueva_fecha_p2.isoformat() if (numero_pagos >= 2 and nuevo_monto_p2 > 0 and nueva_fecha_p2) else None
                                        f3_txt = nueva_fecha_p3.isoformat() if (numero_pagos == 3 and nuevo_monto_p3 > 0 and nueva_fecha_p3) else None

                                        cursor.execute(
                                            """
                                            UPDATE servicios 
                                            SET producto_id = ?, precio = ?, abonos = ?, 
                                                monto_pago_1 = ?, monto_pago_2 = ?, monto_pago_3 = ?, 
                                                fecha_pago_1 = ?, fecha_pago_2 = ?, fecha_pago_3 = ?, 
                                                observaciones = ?
                                            WHERE id = ?
                                            """,
                                            (
                                                nuevo_producto_id, nuevo_precio, nuevos_abonos,
                                                nuevo_monto_p1, nuevo_monto_p2, nuevo_monto_p3,
                                                f1_txt, f2_txt, f3_txt,
                                                nuevas_obs.strip(), servicio_id
                                            )
                                        )

                                        for curso_id, estado in nuevos_estados.items():
                                            actualizar_estado_curso(servicio_id, curso_id, estado)

                                        conexion.commit()
                                        st.session_state.servicio_editando = None
                                        st.success("✅ Cambios guardados correctamente.")
                                        st.rerun()

                        if st.session_state.servicio_eliminando == servicio_id:
                            if saldo > 0:
                                st.error("🚫 No se puede eliminar este trabajo porque tiene valores pendientes.")
                                st.info("💡 Si creaste este registro por error, primero edítalo y pon su precio en 0 para poder borrarlo.")
                            else:
                                confirmar = st.checkbox("Confirmo que deseo eliminar definitivamente este servicio.", key=f"conf_del_{servicio_id}")
                                if confirmar:
                                    if st.button("🗑️ Confirmar eliminación", key=f"exec_del_{servicio_id}"):
                                        eliminar_estados(servicio_id)
                                        cursor.execute("DELETE FROM servicios WHERE id = ?", (servicio_id,))
                                        conexion.commit()
                                        st.session_state.servicio_eliminando = None
                                        st.success("✅ Trabajo eliminado correctamente.")
                                        st.rerun()

                    st.write("")
                st.divider()
            else:
                st.divider()


# ============================================================
# ADMINISTRACIÓN DE USUARIOS
# ============================================================

elif opcion == "👥 Administrar usuarios":
    st.header("👥 Administración de usuarios")
    st.write("Gestiona los usuarios que pueden acceder al sistema.")

    # ====== Cambiar mi propia clave ======
    with st.expander("🔑 Cambiar mi contraseña", expanded=False):
        with st.form("form_cambiar_clave"):
            if es_admin():
                st.info("Eres administrador, puedes cambiar tu clave sin verificar la anterior.")
                clave_actual = ""
            else:
                clave_actual = st.text_input("Clave actual", type="password")

            nueva_clave = st.text_input("Nueva contraseña", type="password")
            confirmar = st.text_input("Confirmar nueva contraseña", type="password")
            submit_clave = st.form_submit_button("Actualizar contraseña")

            if submit_clave:
                if not nueva_clave or not confirmar:
                    st.error("Debes ingresar y confirmar la nueva contraseña.")
                elif nueva_clave != confirmar:
                    st.error("Las nuevas contraseñas no coinciden.")
                elif len(nueva_clave) < 4:
                    st.error("La nueva contraseña debe tener al menos 4 caracteres.")
                else:
                    if es_admin():
                        ok, msg = cambiar_clave(st.session_state.user_id, nueva_clave)
                        if ok:
                            st.success(f"✅ {msg}")
                        else:
                            st.error(f"❌ {msg}")
                    else:
                        if verificar_clave(st.session_state.user_id, clave_actual):
                            ok, msg = cambiar_clave(st.session_state.user_id, nueva_clave)
                            if ok:
                                st.success(f"✅ {msg}")
                            else:
                                st.error(f"❌ {msg}")
                        else:
                            st.error("❌ La clave actual es incorrecta.")

    st.divider()

    # Solo admin puede crear y gestionar usuarios
    if es_admin():
        st.subheader("➕ Crear nuevo usuario")
        with st.form("form_crear_usuario"):
            col1, col2 = st.columns(2)
            with col1:
                nuevo_usuario = st.text_input("Usuario")
                nueva_clave = st.text_input("Clave", type="password")
            with col2:
                nombre = st.text_input("Nombre completo (opcional)")
                rol = st.selectbox("Rol", ["usuario", "admin"], index=0)
            activo = st.checkbox("Activo", value=True)
            submit_crear = st.form_submit_button("Crear usuario")
            if submit_crear:
                if nuevo_usuario and nueva_clave:
                    if len(nueva_clave) < 4:
                        st.error("La clave debe tener al menos 4 caracteres.")
                    else:
                        ok, msg = crear_usuario(nuevo_usuario, nueva_clave, nombre, rol, 1 if activo else 0)
                        if ok:
                            st.success(msg)
                            st.rerun()
                        else:
                            st.error(msg)
                else:
                    st.warning("Usuario y clave son obligatorios.")

        st.divider()
        st.subheader("📋 Lista de usuarios")

        usuarios = obtener_usuarios()
        if not usuarios:
            st.info("No hay usuarios registrados.")
        else:
            # Mostrar en tabla con clave en texto plano (solo admin)
            datos = []
            for u in usuarios:
                datos.append({
                    "ID": u[0],
                    "Usuario": u[1],
                    "Nombre": u[2] or "",
                    "Rol": u[3],
                    "Activo": "✅" if u[4] else "❌",
                    "Clave (texto)": u[5] if u[5] else "—"
                })
            st.dataframe(datos, use_container_width=True)

            # ----- Editar usuario -----
            st.write("### ✏️ Editar usuario")
            usuarios_lista = [f"{u[1]} (ID: {u[0]})" for u in usuarios]
            opcion_usuario = st.selectbox("Seleccionar usuario", usuarios_lista, key="select_usuario_edit")
            id_seleccionado = int(opcion_usuario.split("ID: ")[1].rstrip(")"))

            cursor.execute("SELECT id, usuario, nombre, rol, activo, clave_texto FROM usuarios WHERE id = ?", (id_seleccionado,))
            user_data = cursor.fetchone()
            if user_data:
                with st.form("form_editar_usuario"):
                    nuevo_usuario_login = st.text_input("Usuario (login)", value=user_data[1])
                    nuevo_nombre = st.text_input("Nombre completo", value=user_data[2] or "")
                    nuevo_rol = st.selectbox("Rol", ["usuario", "admin"], index=0 if user_data[3] == "usuario" else 1)
                    nuevo_activo = st.checkbox("Activo", value=bool(user_data[4]))
                    st.text_input("Clave actual (texto plano)", value=user_data[5] or "", disabled=True)
                    submit_editar = st.form_submit_button("💾 Guardar cambios")

                    if submit_editar:
                        cursor.execute(
                            "SELECT id FROM usuarios WHERE usuario = ? AND id != ?",
                            (nuevo_usuario_login.strip(), id_seleccionado)
                        )
                        if cursor.fetchone():
                            st.error("⚠️ El nombre de usuario ya está en uso.")
                        else:
                            if actualizar_usuario(
                                id_seleccionado,
                                usuario=nuevo_usuario_login,
                                nombre=nuevo_nombre,
                                rol=nuevo_rol,
                                activo=nuevo_activo
                            ):
                                st.success("✅ Usuario actualizado correctamente.")
                                st.rerun()
                            else:
                                st.error("❌ Error al actualizar el usuario.")

                # ----- Cambiar clave de este usuario (admin) -----
                with st.expander("🔑 Cambiar clave de este usuario", expanded=False):
                    with st.form("form_cambiar_clave_otro"):
                        nueva_clave_otro = st.text_input("Nueva contraseña", type="password")
                        confirmar_otro = st.text_input("Confirmar", type="password")
                        submit_clave_otro = st.form_submit_button("Actualizar clave")
                        if submit_clave_otro:
                            if nueva_clave_otro and nueva_clave_otro == confirmar_otro:
                                if len(nueva_clave_otro) < 4:
                                    st.error("La contraseña debe tener al menos 4 caracteres.")
                                else:
                                    ok, msg = cambiar_clave(id_seleccionado, nueva_clave_otro)
                                    if ok:
                                        st.success(f"✅ {msg}")
                                        st.rerun()
                                    else:
                                        st.error(f"❌ {msg}")
                            else:
                                st.error("Las contraseñas no coinciden o están vacías.")


# ============================================================
# CERRAR CONEXIÓN
# ============================================================

conexion.close()